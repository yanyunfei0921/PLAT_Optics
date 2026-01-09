import os
import sys
import json
import io
from datetime import datetime
from typing import final
from flask import Flask, render_template, request, jsonify, send_file, send_from_directory
from flask_socketio import SocketIO, emit
sys.path.append(os.getenv('MVCAM_COMMON_RUNENV') + "/Samples/python/MvImport")
from MvCameraControl_class import MvCamera  # type: ignore
from core.cameraService import CameraService, VirtualCameraService
from core.sdiService import SDICameraService, SDI_AVAILABLE
from core.commandService import command_service
from core.databaseService import db_service
serial_service = command_service._serial
app = Flask(
    __name__,
    template_folder = rf"{os.getcwd()}/templates",
    static_folder = rf"{os.getcwd()}/static",
    static_url_path = '/static'
)

app.jinja_env.autoreload = True
app.jinja_env.variable_start_string = '?? --HUA-- <<'
app.jinja_env.variable_end_string = '>> --HUA-- ??'

STATIC_DIR = rf'{os.getcwd()}/static'
CONFIG_DIR = rf'{os.getcwd()}/config'

# 加载应用配置
APP_CONFIG_PATH = os.path.join(CONFIG_DIR, 'app_config.json')
_app_config = {}
try:
    if os.path.exists(APP_CONFIG_PATH):
        with open(APP_CONFIG_PATH, 'r', encoding='utf-8') as f:
            _app_config = json.load(f)
except Exception as e:
    print(f"Warning: Failed to load app_config.json: {e}")

# 从配置文件获取服务器设置，提供默认值
SERVER_HOST = _app_config.get('server', {}).get('host', '127.0.0.1')
SERVER_PORT = _app_config.get('server', {}).get('port', 8090)
SERVER_DEBUG = _app_config.get('server', {}).get('debug', True)

# ========================加载相机配置===============================
def _load_camera_config():
    """加载相机配置文件"""
    config_path = os.path.join(CONFIG_DIR, 'cameraConfig.json')
    try:
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"Warning: Failed to load cameraConfig.json: {e}")
    return {"cameras": []}

def _get_camera_ip(camera_id: int) -> str:
    """从配置文件获取相机IP地址"""
    config = _load_camera_config()
    for cam in config.get('cameras', []):
        if cam.get('id') == camera_id and cam.get('type') == 'mvcamera':
            return cam.get('ip', '')
    return ''

# ========================Socket.IO 初始化===============================
socketio = SocketIO(app, cors_allowed_origins='*')

# 记录每个客户端的推流线程与相机选择
_client_stream_threads = {}
_client_camera_ids = {}

# 初始化测试箱内相机SDK
try:
    ret = MvCamera.MV_CC_Initialize()
    if ret != 0:
        print(f"SDK Init failed! ret[0x{ret:x}]")
    else:
        print("Camera SDK init success")
except Exception as e:
    print(f"SDK init Error: {e}")

# 测试箱内相机实例 (相机1和2为MvCamera硬件)
camSer1 = CameraService(0)
camSer2 = CameraService(1)
# SDI采集相机实例 (相机3为SDI输入)
sdiCam = SDICameraService(camera_id=3)
# 虚拟相机实例 (相机4为静态图像上传模式)
virtualCam = VirtualCameraService(camera_id=4)


def _get_camera_by_id(camera_id):
    """
    根据相机ID获取相机服务实例

    Args:
        camera_id:
            - 1, 2: MvCamera真实相机
            - 3: SDI采集相机
            - 4: 虚拟相机（静态图像上传）

    Returns:
        CameraService, SDICameraService 或 VirtualCameraService 实例
    """
    try:
        cam_id_int = int(camera_id)
    except Exception:
        cam_id_int = 1

    if cam_id_int == 1:
        return camSer1
    elif cam_id_int == 2:
        return camSer2
    elif cam_id_int == 3:
        return sdiCam
    elif cam_id_int == 4:
        return virtualCam
    else:
        return camSer1

def _stream_frames_to_client(camera_service, sid):
    while camera_service.running:
        frame_data = camera_service.getLatestFrame()
        if frame_data:
            socketio.emit('camera_frame', frame_data, room=sid)
        socketio.sleep(0.03)

# OK相机实例


if not os.path.exists(STATIC_DIR):
    os.mkdir(STATIC_DIR)
if not os.path.exists(CONFIG_DIR):
    os.mkdir(CONFIG_DIR)

# ========================页面加载路由==============================================

@app.route('/', methods = ["GET", "POST"])
def index():
    return render_template("index.html")

@app.route('/login', methods = ["GET", "POST"])
def login():
    return render_template("login.html")

@app.route('/serialSetting', methods=["GET"])
def serial_setting():
    return render_template("serialSetting.html")

@app.route('/opticalAxis', methods=["GET"])
def optical_axis():
    return render_template("opticalAxis.html")

@app.route('/laserDistance', methods=["GET"])
def laser_distance():
    return render_template("laserDistance.html")

@app.route('/laserEnergy', methods=["GET"])
def laser_energy():
    return render_template("laserEnergy.html")

@app.route('/dynamicTarget', methods=["GET"])
def dynamic_target():
    return render_template("dynamicTarget.html")

@app.route('/testRecord', methods=["GET"])
def test_record():
    return render_template("testRecord.html")

# 静态文件路由 - 用于访问保存的测试图片
@app.route('/data/<path:filename>')
def serve_data_file(filename):
    """提供data目录下的文件访问"""
    data_dir = os.path.join(os.getcwd(), 'data')
    return send_from_directory(data_dir, filename)

# =======================================================================

@app.route('/get-config', methods = ["GET", "POST"])
def get_config():
    """获取配置信息"""
    pass

# =========================serialSetting页面api==============================================
@app.route('/api/serial-config', methods=['GET'])
def get_serial_config():
    """获取串口配置"""
    config_path = os.path.join(CONFIG_DIR, 'serialConfig.json')
    try:
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            return jsonify({'success': True, 'data': config})
        else:
            return jsonify({'success': False, 'message': '配置文件不存在'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/serial-config', methods=['POST'])
def save_serial_config():
    """保存串口配置"""
    config_path = os.path.join(CONFIG_DIR, 'serialConfig.json')
    try:
        data = request.get_json()
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return jsonify({'success': True, 'message': '配置保存成功'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/serial/connect', methods=['POST'])
def connect_serial():
    """连接串口"""
    try:
        data = request.get_json()
        device_key = data.get('deviceKey')
        port = data.get('port')
        baudrate = data.get('baudrate', 9600)
        data_bits = data.get('dataBits', 8)
        stop_bits = data.get('stopBits', 1)
        parity = data.get('parity', 'None')
        timeout = data.get('timeout', 1000)
        
        success, message = serial_service.connect(
            device_key=device_key,
            port=port,
            baudrate=baudrate,
            data_bits=data_bits,
            stop_bits=stop_bits,
            parity=parity,
            timeout=timeout
        )
        
        return jsonify({'success': success, 'message': message})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/serial/disconnect', methods=['POST'])
def disconnect_serial():
    """断开串口连接"""
    try:
        data = request.get_json()
        device_key = data.get('deviceKey')
        
        success, message = serial_service.disconnect(device_key)
        
        return jsonify({'success': success, 'message': message})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/serial/connect-all', methods=['POST'])
def connect_all_serial():
    """连接所有设备"""
    try:
        devices_config = request.get_json()
        if not isinstance(devices_config, list):
            return jsonify({'success': False, 'message': '请发送设备配置列表'})
        
        # 调用串口服务的批量连接方法
        results = serial_service.connect_all(devices_config)
        
        return jsonify({'success': True, 'results': results})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/serial/disconnect-all', methods=['POST'])
def disconnect_all_serial():
    """断开所有串口连接"""
    try:
        results = serial_service.disconnect_all()
        return jsonify({'success': True, 'results': results})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})



@app.route('/api/serial/send-command', methods=['POST'])
def send_serial_command_raw():
    """发送指令到设备"""
    try:
        data = request.get_json()
        device_key = data.get('deviceKey')
        command = data.get('command')
        wait_response = data.get('waitResponse', True)
        response_timeout = data.get('responseTimeout', 2.0)
        
        # 发送二进制指令
        success, message, response_bytes = serial_service.send_command(
            device_key=device_key,
            command=command,  # 现在是字节串
            wait_response=wait_response,
            response_timeout=response_timeout
        )
        
        return jsonify({
            'success': success,
            'message': message,
            'response': response_bytes.hex(),
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/serial/status', methods=['POST'])
def get_connection_status():
    """获取所有设备的连接状态"""
    try:
        devices_key = request.get_json()
        if not isinstance(devices_key, list):
            return jsonify({'success': False, 'message': '请发送设备键名列表'})
        
        status = serial_service.get_all_connection_status(devices_key)
        return jsonify({'success': True, 'status': status})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
# =========================================================================================

# =========================cameraService api - camearConfig wr part========================
@app.route('/api/camera-config', methods=['GET'])
def get_camera_config():
    """获取相机配置"""
    config_path = os.path.join(CONFIG_DIR, 'cameraConfig.json')
    try:
        if not os.path.exists(config_path):
            # 默认配置
            default_config = {
                "cameras": [
                    {"id": 1, "name": "相机1", "ip": "192.168.1.10", "exposureTime": 5000, "frameRate": 30, "threshold": 128},
                    {"id": 2, "name": "相机2", "ip": "192.168.1.11", "exposureTime": 5000, "frameRate": 30, "threshold": 128}
                ]
            }
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(default_config, f, ensure_ascii=False, indent=2)
            return jsonify({'success': True, 'data': default_config})
        
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        return jsonify({'success': True, 'data': config})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/camera-config', methods=['POST'])
def save_camera_config():
    """保存相机配置"""
    config_path = os.path.join(CONFIG_DIR, 'cameraConfig.json')
    try:
        data = request.get_json()
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return jsonify({'success': True, 'message': '配置保存成功'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
# =========================================================================================

# =========================virtualCamera api (静态图像上传 - 相机4)===========================
@app.route('/api/virtual-camera/upload', methods=['POST'])
def upload_virtual_camera_image():
    """
    上传静态图像到虚拟相机(相机4)进行分析

    接受multipart/form-data格式的图像文件
    可选参数: threshold (二值化阈值), medianKernelSize (中值滤波核大小)
    """
    try:
        if 'image' not in request.files:
            return jsonify({'success': False, 'message': '未找到图像文件'})

        file = request.files['image']
        if file.filename == '':
            return jsonify({'success': False, 'message': '未选择文件'})

        # 读取图像数据
        image_data = file.read()
        filename = file.filename

        # 更新虚拟相机参数（如果前端传递了参数）
        if 'threshold' in request.form:
            try:
                virtualCam.threshold = int(request.form['threshold'])
            except ValueError:
                pass
        if 'medianKernelSize' in request.form:
            try:
                virtualCam.median_kernel_size = int(request.form['medianKernelSize'])
            except ValueError:
                pass

        # 使用虚拟相机处理图像
        result = virtualCam.uploadImage(image_data, filename)

        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/virtual-camera/reprocess', methods=['POST'])
def reprocess_virtual_camera_image():
    """重新处理虚拟相机当前图像 (参数变更后调用)"""
    try:
        result = virtualCam.reprocessImage()
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/virtual-camera/frame', methods=['GET'])
def get_virtual_camera_frame():
    """获取虚拟相机最新帧"""
    try:
        frame_data = virtualCam.getLatestFrame()
        if frame_data:
            return jsonify({'success': True, 'frameData': frame_data})
        else:
            return jsonify({'success': False, 'message': '没有可用帧'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# =========================================================================================

# =========================SDI Camera api (相机3)============================================
@app.route('/api/sdi-camera/devices', methods=['GET'])
def get_sdi_devices():
    """获取SDI设备列表"""
    try:
        if not SDI_AVAILABLE:
            return jsonify({'success': False, 'message': 'SDI SDK不可用'})

        if not sdiCam._initialized:
            success, msg = sdiCam.initialize()
            if not success:
                return jsonify({'success': False, 'message': msg})

        devices = sdiCam.get_devices()
        return jsonify({'success': True, 'devices': devices})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/sdi-camera/params', methods=['GET'])
def get_sdi_params():
    """获取SDI相机当前参数"""
    try:
        params = sdiCam.getAllParams()
        return jsonify({'success': True, 'params': params})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/sdi-camera/params', methods=['POST'])
def set_sdi_params():
    """设置SDI相机参数"""
    try:
        data = request.get_json()
        results = {}

        if 'brightness' in data:
            results['brightness'] = sdiCam.setBrightness(int(data['brightness']))
        if 'contrast' in data:
            results['contrast'] = sdiCam.setContrast(int(data['contrast']))
        if 'saturation' in data:
            results['saturation'] = sdiCam.setSaturation(int(data['saturation']))
        if 'hue' in data:
            results['hue'] = sdiCam.setHue(int(data['hue']))
        if 'threshold' in data:
            results['threshold'] = sdiCam.setThreshold(int(data['threshold']))

        return jsonify({'success': True, 'results': results})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# =========================高级相机参数 API================================================
@app.route('/api/camera/all-params', methods=['POST'])
def get_all_camera_params():
    """获取相机所有参数"""
    try:
        data = request.get_json()
        camera_id = data.get('cameraId', 1)
        cam = _get_camera_by_id(camera_id)

        # 虚拟相机只返回基本参数
        if isinstance(cam, VirtualCameraService):
            params = {
                'threshold': cam.getThreshold(),
                'cameraType': 'virtual'
            }
            return jsonify({'success': True, 'params': params})

        # SDI相机返回SDI特有参数
        if isinstance(cam, SDICameraService):
            params = cam.getAllParams()
            params['cameraType'] = 'sdi'
            return jsonify({'success': True, 'params': params})

        # MvCamera真实相机需要连接后才能获取参数
        if cam.cam is None:
            return jsonify({'success': False, 'message': '相机未连接'})

        params = cam.getAllParams()
        params['cameraType'] = 'mvcamera'
        return jsonify({'success': True, 'params': params})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# =========================================================================================

# =========================commandService api==============================================
@app.route('/api/command/send', methods=['POST'])
def send_command():
    """发送指令"""
    try:
        data = request.get_json()
        device = data.get('device')
        cmd = data.get('cmd')
        params = data.get('params')
        wait_response = data.get('wait_response', True)
        timeout = data.get('timeout', 2.0)
        success, message, response_bytes, decode_result = command_service.send(
            device, cmd, params, wait_response=wait_response, timeout=timeout
        )
        result = {
            'success': success,
            'message': message,
            'response': response_bytes.hex() if response_bytes else ''
        }
        # 如果有解码结果，添加到返回值
        if decode_result is not None:
            result['decode_result'] = decode_result
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/command/get-all-commands', methods=['GET'])
def get_all_commands():
    """获取所有指令"""
    commands = command_service.list_commands()  # 返回的是一个Dict
    return jsonify({'success': True, 'commands': commands})

@app.route('/api/command/config', methods=['GET'])
def get_command_config():
    """获取完整的指令配置（包含模板）"""
    try:
        config = command_service._load_config()
        return jsonify(config)
    except Exception as e:
        return jsonify({'error': str(e)})

# ==================================================================================

# =========================测试记录 API (串口日志 + 光轴测试)===========================

# --------------------- 串口日志 API ---------------------
@app.route('/api/logs/serial', methods=['GET'])
def get_serial_logs():
    """
    查询串口日志
    查询参数:
        - device: 设备筛选
        - port: 串口筛选
        - start_time: 开始时间 (ISO格式)
        - end_time: 结束时间 (ISO格式)
        - limit: 返回数量限制 (默认100)
        - offset: 偏移量 (默认0)
    """
    try:
        device = request.args.get('device')
        port = request.args.get('port')
        start_time = request.args.get('start_time')
        end_time = request.args.get('end_time')
        limit = int(request.args.get('limit', 100))
        offset = int(request.args.get('offset', 0))

        result = db_service.get_serial_logs(
            device=device,
            port=port,
            start_time=start_time,
            end_time=end_time,
            limit=limit,
            offset=offset
        )
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/logs/serial/<int:log_id>', methods=['DELETE'])
def delete_serial_log(log_id):
    """删除单条串口日志"""
    try:
        success = db_service.delete_serial_log(log_id)
        if success:
            return jsonify({'success': True, 'message': '删除成功'})
        else:
            return jsonify({'success': False, 'message': '记录不存在'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/logs/serial/clear', methods=['POST'])
def clear_serial_logs():
    """清空所有串口日志"""
    try:
        count = db_service.clear_serial_logs()
        return jsonify({'success': True, 'message': f'已清空 {count} 条记录'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# --------------------- 光轴测试记录 API ---------------------
@app.route('/api/tests/optical-axis', methods=['GET'])
def get_optical_tests():
    """
    查询光轴测试记录
    查询参数:
        - start_time: 开始时间 (ISO格式)
        - end_time: 结束时间 (ISO格式)
        - limit: 返回数量限制 (默认50)
        - offset: 偏移量 (默认0)
    """
    try:
        start_time = request.args.get('start_time')
        end_time = request.args.get('end_time')
        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))

        result = db_service.get_optical_tests(
            start_time=start_time,
            end_time=end_time,
            limit=limit,
            offset=offset
        )
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/tests/optical-axis/<int:test_id>', methods=['GET'])
def get_optical_test(test_id):
    """获取单条光轴测试记录"""
    try:
        record = db_service.get_optical_test(test_id)
        if record:
            return jsonify({'success': True, 'record': record})
        else:
            return jsonify({'success': False, 'message': '记录不存在'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/tests/optical-axis', methods=['POST'])
def save_optical_test():
    """
    保存光轴测试记录 (创建基准光轴记录)

    请求体JSON:
        - operator: 操作人员
        - base_camera_id: 基准相机ID
        - base_camera_name: 基准相机名称
        - base_image: 基准图片base64数据
        - base_width: 基准图像宽度
        - base_height: 基准图像高度
        - base_centroid_x: 基准质心X
        - base_centroid_y: 基准质心Y
        - base_focal_length: 基准焦距(mm)
        - base_pixel_size: 基准像元大小(μm)
        - base_offset_x: 基准偏移X(deg)
        - base_offset_y: 基准偏移Y(deg)
    """
    try:
        data = request.get_json()

        # 保存基准图片
        base_image_path = None
        if data.get('base_image'):
            base_image_path = db_service.save_image_base64(data['base_image'], 'base')

        record_data = {
            'operator': data.get('operator'),
            'base_camera_id': data.get('base_camera_id'),
            'base_camera_name': data.get('base_camera_name'),
            'base_image_path': base_image_path,
            'base_width': data.get('base_width'),
            'base_height': data.get('base_height'),
            'base_centroid_x': data.get('base_centroid_x'),
            'base_centroid_y': data.get('base_centroid_y'),
            'base_focal_length': data.get('base_focal_length'),
            'base_pixel_size': data.get('base_pixel_size'),
            'base_offset_x': data.get('base_offset_x'),
            'base_offset_y': data.get('base_offset_y'),
        }

        record_id = db_service.save_optical_test(record_data)
        return jsonify({'success': True, 'id': record_id, 'message': '基准光轴记录已保存'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/tests/optical-axis/<int:test_id>', methods=['PUT'])
def update_optical_test(test_id):
    """
    更新光轴测试记录 (补充测试光轴数据或备注)

    请求体JSON:
        - test_camera_id: 测试相机ID
        - test_camera_name: 测试相机名称
        - test_image: 测试图片base64数据
        - test_width: 测试图像宽度
        - test_height: 测试图像高度
        - test_centroid_x: 测试质心X
        - test_centroid_y: 测试质心Y
        - test_focal_length: 测试焦距(mm)
        - test_pixel_size: 测试像元大小(μm)
        - test_offset_x: 测试偏移X(deg)
        - test_offset_y: 测试偏移Y(deg)
        - remark: 备注
    """
    try:
        data = request.get_json()
        update_data = {}

        # 保存测试图片
        if data.get('test_image'):
            test_image_path = db_service.save_image_base64(data['test_image'], 'test')
            update_data['test_image_path'] = test_image_path

        # 收集测试光轴数据
        for field in ['test_camera_id', 'test_camera_name', 'test_width', 'test_height',
                      'test_centroid_x', 'test_centroid_y', 'test_focal_length',
                      'test_pixel_size', 'test_offset_x', 'test_offset_y', 'remark']:
            if field in data:
                update_data[field] = data[field]

        if not update_data:
            return jsonify({'success': False, 'message': '没有要更新的数据'})

        success = db_service.update_optical_test(test_id, update_data)
        if success:
            return jsonify({'success': True, 'message': '测试记录已更新'})
        else:
            return jsonify({'success': False, 'message': '记录不存在或更新失败'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/tests/optical-axis/<int:test_id>', methods=['DELETE'])
def delete_optical_test(test_id):
    """删除光轴测试记录（同时删除关联图片）"""
    try:
        success = db_service.delete_optical_test(test_id)
        if success:
            return jsonify({'success': True, 'message': '删除成功'})
        else:
            return jsonify({'success': False, 'message': '记录不存在'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# --------------------- 导出 API ---------------------
def hex_to_string(hex_str):
    """将HEX字符串转换为可读字符串"""
    if not hex_str or hex_str == '-':
        return ''
    try:
        # 移除空格，将HEX字符串按每2个字符分割
        hex_clean = hex_str.replace(' ', '')
        result = ''
        for i in range(0, len(hex_clean), 2):
            if i + 2 <= len(hex_clean):
                hex_byte = hex_clean[i:i+2]
                char_code = int(hex_byte, 16)
                # 只显示可打印字符(32-126)和常见扩展ASCII(128-255)，其他用点号表示
                if (32 <= char_code <= 126) or (128 <= char_code <= 255):
                    result += chr(char_code)
                else:
                    result += '.'
        return result
    except Exception:
        return hex_str  # 转换失败返回原字符串


@app.route('/api/export/serial-logs/excel', methods=['GET'])
def export_serial_logs_excel():
    """导出串口日志为Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side

        # 获取所有日志
        result = db_service.get_serial_logs(limit=1000, offset=0)
        logs = result['logs']

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '串口日志'

        # 设置表头
        headers = ['ID', '时间', '设备', '串口', '指令', '参数',
                   '发送数据(HEX)', '发送数据(字符)', '返回数据(HEX)', '返回数据(字符)',
                   '状态', '消息']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 填充数据
        for row, log in enumerate(logs, 2):
            ws.cell(row=row, column=1, value=log['id'])
            # UTC时间转换为本地时间
            timestamp = log['timestamp']
            if timestamp:
                try:
                    utc_str = str(timestamp).replace(' ', 'T')
                    if not utc_str.endswith('Z') and '+' not in utc_str:
                        utc_str += '+00:00'
                    utc_dt = datetime.fromisoformat(utc_str.replace('Z', '+00:00'))
                    local_dt = utc_dt.astimezone()
                    timestamp = local_dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    pass
            ws.cell(row=row, column=2, value=timestamp)
            ws.cell(row=row, column=3, value=log['device'])
            ws.cell(row=row, column=4, value=log['port'])
            ws.cell(row=row, column=5, value=log['command'])
            ws.cell(row=row, column=6, value=log['params'])
            # 发送数据 HEX
            ws.cell(row=row, column=7, value=log['cmd_bytes'])
            # 发送数据 字符
            ws.cell(row=row, column=8, value=hex_to_string(log['cmd_bytes']))
            # 返回数据 HEX
            ws.cell(row=row, column=9, value=log['response_bytes'])
            # 返回数据 字符
            ws.cell(row=row, column=10, value=hex_to_string(log['response_bytes']))
            # 状态和消息
            ws.cell(row=row, column=11, value='成功' if log['success'] else '失败')
            ws.cell(row=row, column=12, value=log['message'])

        # 调整列宽
        column_widths = [8, 20, 15, 10, 20, 30, 30, 25, 30, 25, 8, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'serial_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=filename)
    except ImportError:
        return jsonify({'success': False, 'message': '请安装openpyxl库: pip install openpyxl'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/export/optical-tests/excel', methods=['GET'])
def export_optical_tests_excel():
    """导出光轴测试记录为Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment

        # 获取所有记录
        result = db_service.get_optical_tests(limit=1000, offset=0)
        records = result['records']

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '光轴测试记录'

        # 设置表头
        headers = ['ID', '测试时间',
                  '基准相机ID', '基准相机', '基准图像路径', '基准宽度', '基准高度',
                  '基准质心X', '基准质心Y', '基准焦距(mm)', '基准像元(μm)',
                  '基准偏移X(°)', '基准偏移Y(°)',
                  '测试相机ID', '测试相机', '测试图像路径', '测试宽度', '测试高度',
                  '测试质心X', '测试质心Y', '测试焦距(mm)', '测试像元(μm)',
                  '测试偏移X(°)', '测试偏移Y(°)', '备注']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 填充数据
        for row, rec in enumerate(records, 2):
            ws.cell(row=row, column=1, value=rec['id'])
            # UTC时间转换为本地时间
            test_time = rec['test_time']
            if test_time:
                try:
                    utc_str = str(test_time).replace(' ', 'T')
                    if not utc_str.endswith('Z') and '+' not in utc_str:
                        utc_str += '+00:00'
                    utc_dt = datetime.fromisoformat(utc_str.replace('Z', '+00:00'))
                    local_dt = utc_dt.astimezone()
                    test_time = local_dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    pass
            ws.cell(row=row, column=2, value=test_time)
            ws.cell(row=row, column=3, value=rec['base_camera_id'])
            ws.cell(row=row, column=4, value=rec['base_camera_name'])
            ws.cell(row=row, column=5, value=rec['base_image_path'])
            ws.cell(row=row, column=6, value=rec['base_width'])
            ws.cell(row=row, column=7, value=rec['base_height'])
            ws.cell(row=row, column=8, value=rec['base_centroid_x'])
            ws.cell(row=row, column=9, value=rec['base_centroid_y'])
            ws.cell(row=row, column=10, value=rec['base_focal_length'])
            ws.cell(row=row, column=11, value=rec['base_pixel_size'])
            ws.cell(row=row, column=12, value=rec['base_offset_x'])
            ws.cell(row=row, column=13, value=rec['base_offset_y'])
            ws.cell(row=row, column=14, value=rec['test_camera_id'])
            ws.cell(row=row, column=15, value=rec['test_camera_name'])
            ws.cell(row=row, column=16, value=rec['test_image_path'])
            ws.cell(row=row, column=17, value=rec['test_width'])
            ws.cell(row=row, column=18, value=rec['test_height'])
            ws.cell(row=row, column=19, value=rec['test_centroid_x'])
            ws.cell(row=row, column=20, value=rec['test_centroid_y'])
            ws.cell(row=row, column=21, value=rec['test_focal_length'])
            ws.cell(row=row, column=22, value=rec['test_pixel_size'])
            ws.cell(row=row, column=23, value=rec['test_offset_x'])
            ws.cell(row=row, column=24, value=rec['test_offset_y'])
            ws.cell(row=row, column=25, value=rec['remark'])

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'optical_tests_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=filename)
    except ImportError:
        return jsonify({'success': False, 'message': '请安装openpyxl库: pip install openpyxl'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/export/optical-test/<int:test_id>/pdf', methods=['GET'])
def export_optical_test_pdf(test_id):
    """导出单条光轴测试记录为PDF报告（单页A4）"""
    try:
        # 解决 reportlab 与 Python/OpenSSL 兼容性问题
        import hashlib
        _orig_md5 = hashlib.md5
        def _patched_md5(*args, **kwargs):
            kwargs.pop('usedforsecurity', None)
            return _orig_md5(*args, **kwargs)
        hashlib.md5 = _patched_md5

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        record = db_service.get_optical_test(test_id)
        if not record:
            return jsonify({'success': False, 'message': '记录不存在'})

        # 尝试注册中文字体
        font_name = 'Helvetica'
        try:
            font_paths = [
                'C:/Windows/Fonts/simhei.ttf',
                'C:/Windows/Fonts/msyh.ttc',
                'simhei.ttf',
            ]
            for fp in font_paths:
                try:
                    pdfmetrics.registerFont(TTFont('SimHei', fp))
                    font_name = 'SimHei'
                    break
                except:
                    continue
        except:
            pass

        def fmt(val, decimals=2):
            if val is None:
                return '-'
            return f'{val:.{decimals}f}'

        # 创建PDF - 紧凑边距
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            topMargin=12*mm,
            bottomMargin=10*mm,
            leftMargin=15*mm,
            rightMargin=15*mm
        )
        elements = []

        # 样式定义 - 紧凑字体
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle', fontName=font_name, fontSize=18,
            textColor=colors.HexColor('#1890ff'), spaceAfter=8, alignment=1
        )
        heading_style = ParagraphStyle(
            'CustomHeading', fontName=font_name, fontSize=11,
            textColor=colors.HexColor('#333333'), spaceBefore=6, spaceAfter=4
        )
        small_style = ParagraphStyle(
            'Small', fontName=font_name, fontSize=9, textColor=colors.HexColor('#666666')
        )

        # 标题
        elements.append(Paragraph('光轴一致性测试报告', title_style))

        # 基本信息 - 单行
        test_time = record['test_time'] or '-'
        if test_time != '-':
            # SQLite CURRENT_TIMESTAMP 存储的是UTC时间，转换为本地时间
            try:
                from datetime import timezone
                utc_str = str(test_time).replace(' ', 'T')
                if not utc_str.endswith('Z') and '+' not in utc_str:
                    utc_str += '+00:00'  # 标记为UTC
                utc_dt = datetime.fromisoformat(utc_str.replace('Z', '+00:00'))
                local_dt = utc_dt.astimezone()  # 转换为本地时间
                test_time = local_dt.strftime('%Y-%m-%d %H:%M:%S')
            except:
                # 解析失败，使用简单格式化
                if 'T' in str(test_time):
                    test_time = str(test_time).replace('T', ' ')[:19]

        info_data = [[
            f"编号: #{record['id']}",
            f"测试时间: {test_time}",
            f"操作人员: {record.get('operator') or '-'}"
        ]]
        info_table = Table(info_data, colWidths=[50*mm, 70*mm, 60*mm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#666666')),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 4*mm))

        # 准备图片
        base_img = None
        test_img = None
        img_width, img_height = 55*mm, 42*mm

        if record.get('base_image_path'):
            try:
                img_path = os.path.join(os.getcwd(), record['base_image_path'])
                if os.path.exists(img_path):
                    base_img = RLImage(img_path, width=img_width, height=img_height)
            except:
                pass

        if record.get('test_image_path'):
            try:
                img_path = os.path.join(os.getcwd(), record['test_image_path'])
                if os.path.exists(img_path):
                    test_img = RLImage(img_path, width=img_width, height=img_height)
            except:
                pass

        # 双栏布局 - 基准光轴 | 测试光轴
        col_width = 85*mm
        row_height = 5.5*mm

        # 表头
        header_data = [['基准光轴', '测试光轴']]
        header_table = Table(header_data, colWidths=[col_width, col_width])
        header_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('TEXTCOLOR', (0, 0), (0, 0), colors.HexColor('#52c41a')),
            ('TEXTCOLOR', (1, 0), (1, 0), colors.HexColor('#fa8c16')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#f6ffed')),
            ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#fff7e6')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#d9d9d9')),
        ]))
        elements.append(header_table)

        # 数据对比表格
        data_rows = [
            ['相机', record['base_camera_name'] or str(record['base_camera_id'] or '-'),
             '相机', record['test_camera_name'] or str(record['test_camera_id'] or '-')],
            ['尺寸', f"{record['base_width'] or '-'}x{record['base_height'] or '-'}",
             '尺寸', f"{record['test_width'] or '-'}x{record['test_height'] or '-'}"],
            ['质心', f"({fmt(record['base_centroid_x'])}, {fmt(record['base_centroid_y'])})",
             '质心', f"({fmt(record['test_centroid_x'])}, {fmt(record['test_centroid_y'])})"],
            ['焦距', f"{fmt(record['base_focal_length'])} mm",
             '焦距', f"{fmt(record['test_focal_length'])} mm"],
            ['像元', f"{fmt(record['base_pixel_size'])} um",
             '像元', f"{fmt(record['test_pixel_size'])} um"],
            ['偏移X', f"{fmt(record['base_offset_x'], 4)}°",
             '偏移X', f"{fmt(record['test_offset_x'], 4)}°"],
            ['偏移Y', f"{fmt(record['base_offset_y'], 4)}°",
             '偏移Y', f"{fmt(record['test_offset_y'], 4)}°"],
        ]

        data_table = Table(data_rows, colWidths=[18*mm, 67*mm, 18*mm, 67*mm], rowHeights=[row_height]*7)
        data_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#888888')),
            ('TEXTCOLOR', (2, 0), (2, -1), colors.HexColor('#888888')),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#333333')),
            ('TEXTCOLOR', (3, 0), (3, -1), colors.HexColor('#333333')),
            ('BACKGROUND', (0, 0), (1, -1), colors.HexColor('#fafafa')),
            ('BACKGROUND', (2, 0), (3, -1), colors.HexColor('#fffbf0')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e8e8e8')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        elements.append(data_table)

        # 图片行
        if base_img or test_img:
            elements.append(Spacer(1, 3*mm))
            img_row = [[base_img or '', test_img or '']]
            img_table = Table(img_row, colWidths=[col_width, col_width])
            img_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOX', (0, 0), (0, 0), 0.5, colors.HexColor('#d9d9d9') if base_img else colors.white),
                ('BOX', (1, 0), (1, 0), 0.5, colors.HexColor('#d9d9d9') if test_img else colors.white),
                ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#fafafa') if base_img else colors.white),
                ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#fafafa') if test_img else colors.white),
            ]))
            elements.append(img_table)

        # 备注
        if record.get('remark'):
            elements.append(Spacer(1, 4*mm))
            remark_data = [['备注', record['remark']]]
            remark_table = Table(remark_data, colWidths=[18*mm, 152*mm])
            remark_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TEXTCOLOR', (0, 0), (0, 0), colors.HexColor('#d48806')),
                ('TEXTCOLOR', (1, 0), (1, 0), colors.HexColor('#666666')),
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fffbe6')),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#ffe58f')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(remark_table)

        # 页脚
        elements.append(Spacer(1, 6*mm))
        footer_data = [[
            '光轴一致性测试系统',
            f"报告生成: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ]]
        footer_table = Table(footer_data, colWidths=[90*mm, 80*mm])
        footer_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#999999')),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ]))
        elements.append(footer_table)

        doc.build(elements)
        output.seek(0)

        filename = f'optical_test_report_{test_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        return send_file(output, mimetype='application/pdf', as_attachment=True, download_name=filename)

    except ImportError as e:
        return jsonify({'success': False, 'message': f'请安装reportlab库: pip install reportlab. 错误: {str(e)}'})
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'message': str(e), 'trace': traceback.format_exc()})

# ==================================================================================


# =========================Socket.IO 事件============================================
@socketio.on('disconnect')
def handle_client_disconnect():
    """
    客户端断开连接时清理资源
    防止_client_stream_threads和_client_camera_ids内存泄漏
    """
    sid = request.sid
    camera_id = _client_camera_ids.pop(sid, None)

    # 清理推流线程记录
    _client_stream_threads.pop(sid, None)

    # 如果该客户端有绑定相机，尝试停止相机采集
    if camera_id is not None:
        try:
            cam = _get_camera_by_id(camera_id)
            if cam is not None and cam.running:
                if isinstance(cam, VirtualCameraService):
                    # 虚拟相机无需清理
                    pass
                elif isinstance(cam, SDICameraService):
                    # SDI相机断开
                    cam.disconnect()
                else:
                    # MvCamera断开
                    cam.closeAndDisconnectCamera()
                print(f"Client {sid} disconnected, camera {camera_id} stopped")
        except Exception as e:
            print(f"Error cleaning up camera {camera_id} for client {sid}: {e}")


@socketio.on('camera_connect')
def handle_camera_connect(data):
    """
    连接并开始采集
    data 需包含 cameraId:
        - 1, 2: MvCamera真实相机
        - 3: SDI采集相机
        - 4: 虚拟相机（静态图像上传）
    """
    try:
        camera_id = data.get('cameraId', 1)
        cam = _get_camera_by_id(camera_id)

        # 虚拟相机（相机4）不需要初始化SDK，直接返回成功
        if isinstance(cam, VirtualCameraService):
            _client_camera_ids[request.sid] = int(camera_id)
            emit('camera_connected', {
                'success': True,
                'cameraId': int(camera_id),
                'cameraType': 'virtual',
                'isVirtualCamera': True,
                'message': '虚拟相机已就绪，请上传图像文件'
            }, room=request.sid)
            return

        # SDI相机（相机3）初始化和连接
        if isinstance(cam, SDICameraService):
            if not SDI_AVAILABLE:
                emit('camera_error', {'success': False, 'message': 'SDI SDK不可用', 'cameraId': int(camera_id)}, room=request.sid)
                return

            sdi_channel = data.get('sdiChannel', 0)
            resolution_index = data.get('resolutionIndex', 0)

            success, msg = cam.connect(channel=sdi_channel, resolution_index=resolution_index)
            if not success:
                emit('camera_error', {'success': False, 'message': msg, 'cameraId': int(camera_id)}, room=request.sid)
                return

            # 记录客户端选择的相机
            _client_camera_ids[request.sid] = int(camera_id)

            # 启动向该客户端推流的后台任务
            def stream_sdi_frames(sdi_cam, sid):
                while sdi_cam.running:
                    frame_data = sdi_cam.getFrame()
                    if frame_data:
                        socketio.emit('camera_frame', frame_data, room=sid)
                    socketio.sleep(0.03)

            if request.sid in _client_stream_threads:
                pass
            _client_stream_threads[request.sid] = socketio.start_background_task(stream_sdi_frames, cam, request.sid)

            emit('camera_connected', {
                'success': True,
                'cameraId': int(camera_id),
                'cameraType': 'sdi',
                'isVirtualCamera': False,
                'message': 'SDI采集已连接'
            }, room=request.sid)
            return

        # MvCamera真实相机（相机1和2）：使用IP地址查找并连接
        camera_ip = _get_camera_ip(int(camera_id))
        if not camera_ip:
            emit('camera_error', {'success': False, 'message': f'未找到相机{camera_id}的IP配置', 'cameraId': int(camera_id)}, room=request.sid)
            return

        # 枚举设备并找到匹配IP的相机
        if not cam.initCameraByIp(camera_ip):
            emit('camera_error', {'success': False, 'message': f'未找到IP为 {camera_ip} 的相机设备', 'cameraId': int(camera_id)}, room=request.sid)
            return
        if not cam.connectAndOpenCamera():
            emit('camera_error', {'success': False, 'message': f'连接或开始采集失败 (IP: {camera_ip})', 'cameraId': int(camera_id)}, room=request.sid)
            return

        # 记录客户端选择的相机
        _client_camera_ids[request.sid] = int(camera_id)

        # 启动向该客户端推流的后台任务
        if request.sid in _client_stream_threads:
            # 已有线程则跳过或覆盖
            pass
        _client_stream_threads[request.sid] = socketio.start_background_task(_stream_frames_to_client, cam, request.sid)

        emit('camera_connected', {
            'success': True,
            'cameraId': int(camera_id),
            'cameraType': 'mvcamera',
            'isVirtualCamera': False
        }, room=request.sid)
    except Exception as e:
        emit('camera_error', {'success': False, 'message': str(e)}, room=request.sid)


@socketio.on('camera_disconnect')
def handle_camera_disconnect(data):
    """停止采集并断开连接，data 可包含 cameraId（可选，不传则用该连接最近一次选择）"""
    try:
        camera_id = data.get('cameraId') if isinstance(data, dict) else None
        if camera_id is None:
            camera_id = _client_camera_ids.get(request.sid)

        cam = _get_camera_by_id(camera_id) if camera_id is not None else None

        # 根据相机类型断开连接
        if cam is not None:
            if isinstance(cam, VirtualCameraService):
                # 虚拟相机不需要断开连接
                pass
            elif isinstance(cam, SDICameraService):
                # SDI相机断开连接
                cam.disconnect()
            else:
                # MvCamera断开连接
                cam.closeAndDisconnectCamera()

        # 清理该客户端记录
        _client_camera_ids.pop(request.sid, None)
        _client_stream_threads.pop(request.sid, None)

        emit('camera_disconnected', {'success': True, 'cameraId': int(camera_id) if camera_id is not None else None}, room=request.sid)
    except Exception as e:
        emit('camera_error', {'success': False, 'message': str(e)}, room=request.sid)

@socketio.on('camera_set_param')
def handle_camera_set_param(data):
    """
    设置相机参数
    data格式: {
        'cameraId': 1,
        'type': 参数类型,
        'value': 参数值
    }

    支持的参数类型:
    - 基础参数: exposureTime, frameRate, gain, threshold, imageMode
    - MvCamera高级参数 (相机1/2):
        - 图像尺寸: width, height, offsetX, offsetY
        - 图像处理: gamma, blackLevel, reverseX, reverseY
        - 触发控制: triggerMode, triggerSource, exposureAuto, gainAuto
    - SDI参数 (相机3):
        - 视频参数: brightness, contrast, saturation, hue
    """
    try:
        camera_id = data.get('cameraId', 1)
        param_type = data.get('type')
        value = data.get('value')

        cam = _get_camera_by_id(camera_id)

        # 虚拟相机（相机4）只支持 threshold、medianKernelSize 和 imageMode
        if isinstance(cam, VirtualCameraService):
            success = False
            message = ""

            if param_type == 'threshold':
                success = cam.setThreshold(value)
                message = "设置二值化阈值成功" if success else "设置二值化阈值失败"
            elif param_type == 'medianKernelSize':
                success = cam.setMedianKernelSize(int(value))
                message = f"设置中值滤波核大小为 {value}" if success else "设置中值滤波核大小失败"
            elif param_type == 'imageMode':
                is_binary = True if int(value) == 1 else False
                success = cam.setReturnBinaryMode(is_binary)
                mode_str = "二值化图" if is_binary else "原始灰度图"
                message = f"已切换至{mode_str}" if success else "切换显示模式失败"
            else:
                message = f"虚拟相机不支持参数: {param_type}"
                success = False

            if success:
                emit('camera_msg', {'success': True, 'message': message}, room=request.sid)
            else:
                emit('camera_error', {'success': False, 'message': message}, room=request.sid)
            return

        # SDI相机（相机3）支持特有参数
        if isinstance(cam, SDICameraService):
            success = False
            message = ""

            if param_type == 'threshold':
                success = cam.setThreshold(int(value))
                message = "设置二值化阈值成功" if success else "设置二值化阈值失败"
            elif param_type == 'medianKernelSize':
                success = cam.setMedianKernelSize(int(value))
                message = f"设置中值滤波核大小为 {value}" if success else "设置中值滤波核大小失败"
            elif param_type == 'imageMode':
                success = cam.setImageMode(int(value))
                mode_str = "二值化图" if int(value) == 1 else "原始图"
                message = f"已切换至{mode_str}" if success else "切换显示模式失败"
            elif param_type == 'brightness':
                success = cam.setBrightness(int(value))
                message = "设置亮度成功" if success else "设置亮度失败"
            elif param_type == 'contrast':
                success = cam.setContrast(int(value))
                message = "设置对比度成功" if success else "设置对比度失败"
            elif param_type == 'saturation':
                success = cam.setSaturation(int(value))
                message = "设置饱和度成功" if success else "设置饱和度失败"
            elif param_type == 'hue':
                success = cam.setHue(int(value))
                message = "设置色调成功" if success else "设置色调失败"
            else:
                message = f"SDI相机不支持参数: {param_type}"
                success = False

            if success:
                emit('camera_msg', {'success': True, 'message': message}, room=request.sid)
            else:
                emit('camera_error', {'success': False, 'message': message}, room=request.sid)
            return

        # MvCamera真实相机（相机1和2）：检查是否已连接
        if cam.cam is None:
            emit('camera_error', {'success': False, 'message': f'相机{camera_id}未连接，无法设置参数'}, room=request.sid)
            return

        success = False
        message = ""

        # 基础参数
        if param_type == 'exposureTime':
            success = cam.setExposureTime(float(value))
            message = "设置曝光时间成功" if success else "设置曝光时间失败"
        elif param_type == 'frameRate':
            success = cam.setAcquisitionFrameRate(float(value))
            message = "设置帧率成功" if success else "设置帧率失败"
        elif param_type == 'gain':
            success = cam.setGain(float(value))
            message = "设置增益成功" if success else "设置增益失败"
        elif param_type == 'threshold':
            success = cam.setThreshold(value)
            message = "设置二值化阈值成功" if success else "设置二值化阈值失败"
        elif param_type == 'medianKernelSize':
            success = cam.setMedianKernelSize(int(value))
            message = f"设置中值滤波核大小为 {value}" if success else "设置中值滤波核大小失败"
        elif param_type == 'imageMode':
            is_binary = True if int(value) == 1 else False
            success = cam.setReturnBinaryMode(is_binary)
            mode_str = "二值化图" if is_binary else "原始灰度图"
            message = f"已切换至{mode_str}" if success else "切换显示模式失败"

        # 高级参数 - 图像尺寸
        elif param_type == 'width':
            success = cam.setWidth(int(value))
            message = "设置宽度成功" if success else "设置宽度失败"
        elif param_type == 'height':
            success = cam.setHeight(int(value))
            message = "设置高度成功" if success else "设置高度失败"
        elif param_type == 'offsetX':
            success = cam.setOffsetX(int(value))
            message = "设置X偏移成功" if success else "设置X偏移失败"
        elif param_type == 'offsetY':
            success = cam.setOffsetY(int(value))
            message = "设置Y偏移成功" if success else "设置Y偏移失败"

        # 高级参数 - 图像处理
        elif param_type == 'gamma':
            success = cam.setGamma(float(value))
            message = "设置Gamma成功" if success else "设置Gamma失败"
        elif param_type == 'blackLevel':
            success = cam.setBlackLevel(float(value))
            message = "设置黑电平成功" if success else "设置黑电平失败"
        elif param_type == 'reverseX':
            success = cam.setReverseX(bool(value))
            message = "设置X镜像成功" if success else "设置X镜像失败"
        elif param_type == 'reverseY':
            success = cam.setReverseY(bool(value))
            message = "设置Y镜像成功" if success else "设置Y镜像失败"

        # 高级参数 - 触发控制
        elif param_type == 'triggerMode':
            success = cam.setTriggerMode(int(value))
            message = "设置触发模式成功" if success else "设置触发模式失败"
        elif param_type == 'triggerSource':
            success = cam.setTriggerSource(int(value))
            message = "设置触发源成功" if success else "设置触发源失败"
        elif param_type == 'exposureAuto':
            success = cam.setExposureAuto(int(value))
            message = "设置自动曝光成功" if success else "设置自动曝光失败"
        elif param_type == 'gainAuto':
            success = cam.setGainAuto(int(value))
            message = "设置自动增益成功" if success else "设置自动增益失败"
        elif param_type == 'softwareTrigger':
            success = cam.softwareTrigger()
            message = "软触发成功" if success else "软触发失败"

        else:
            message = f"未知参数类型: {param_type}"
            success = False

        if success:
            emit('camera_msg', {'success': True, 'message': message}, room=request.sid)
        else:
            emit('camera_error', {'success': False, 'message': message}, room=request.sid)

    except Exception as e:
        emit('camera_error', {'success': False, 'message': str(e)}, room=request.sid)


@socketio.on('camera_get_param')
def handle_camera_get_param(data):
    """
    获取相机参数
    data格式: {
        'cameraId': 1,
        'type': 参数类型
    }

    支持与 camera_set_param 相同的参数类型
    """
    try:
        camera_id = data.get('cameraId', 1)
        param_type = data.get('type')

        cam = _get_camera_by_id(camera_id)
        is_virtual = isinstance(cam, VirtualCameraService)

        # 虚拟相机只支持 threshold
        if is_virtual:
            if param_type == 'threshold':
                value = cam.getThreshold()
                emit('camera_param_value', {'success': True, 'type': param_type, 'value': value, 'cameraId': camera_id}, room=request.sid)
            else:
                emit('camera_error', {'success': False, 'message': f'虚拟相机不支持参数: {param_type}'}, room=request.sid)
            return

        # threshold 不需要相机连接
        if param_type == 'threshold':
            value = cam.getThreshold()
            emit('camera_param_value', {'success': True, 'type': param_type, 'value': value, 'cameraId': camera_id}, room=request.sid)
            return

        # 其他参数需要调用SDK，必须确保相机已连接
        if cam.cam is None:
            emit('camera_error', {'success': False, 'message': f'相机{camera_id}未连接，无法获取参数'}, room=request.sid)
            return

        value = None

        # 基础参数
        if param_type == 'exposureTime':
            value = cam.getExposureTime()
        elif param_type == 'frameRate':
            value = cam.getAcquisitionFrameRate()
        elif param_type == 'gain':
            value = cam.getGain()

        # 高级参数 - 图像尺寸
        elif param_type == 'width':
            value = cam.getWidth()
        elif param_type == 'height':
            value = cam.getHeight()
        elif param_type == 'offsetX':
            value = cam.getOffsetX()
        elif param_type == 'offsetY':
            value = cam.getOffsetY()

        # 高级参数 - 图像处理
        elif param_type == 'gamma':
            value = cam.getGamma()
        elif param_type == 'blackLevel':
            value = cam.getBlackLevel()
        elif param_type == 'reverseX':
            value = cam.getReverseX()
        elif param_type == 'reverseY':
            value = cam.getReverseY()

        # 高级参数 - 触发控制
        elif param_type == 'triggerMode':
            value = cam.getTriggerMode()
        elif param_type == 'triggerSource':
            value = cam.getTriggerSource()
        elif param_type == 'exposureAuto':
            value = cam.getExposureAuto()
        elif param_type == 'gainAuto':
            value = cam.getGainAuto()

        # 获取所有参数
        elif param_type == 'all':
            params = cam.getAllParams()
            emit('camera_param_value', {'success': True, 'type': 'all', 'value': params, 'cameraId': camera_id}, room=request.sid)
            return

        else:
            emit('camera_error', {'success': False, 'message': f"未知参数类型: {param_type}"}, room=request.sid)
            return

        if value is not None and value != -1:
            emit('camera_param_value', {'success': True, 'type': param_type, 'value': value, 'cameraId': camera_id}, room=request.sid)
        else:
            emit('camera_error', {'success': False, 'message': f"获取{param_type}失败"}, room=request.sid)

    except Exception as e:
        emit('camera_error', {'success': False, 'message': str(e)}, room=request.sid)

# ============================ main function ========================================
if __name__ == "__main__":
    try:
        print(f"Starting server at http://{SERVER_HOST}:{SERVER_PORT}")
        socketio.run(app, host=SERVER_HOST, port=SERVER_PORT, debug=SERVER_DEBUG)
    finally:
        MvCamera.MV_CC_Finalize()